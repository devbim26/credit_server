"""Aggregation + OpenRouter account polling + sanitized settings for the dashboard."""
from __future__ import annotations

import asyncio
import json
import logging
import time
from datetime import datetime, timedelta, timezone
from typing import Any

import httpx

from . import db
from .settings import settings

log = logging.getLogger("analytics.stats")

_RANGE_DELTAS = {"1h": timedelta(hours=1), "24h": timedelta(hours=24),
                 "7d": timedelta(days=7), "30d": timedelta(days=30)}
VALID_RANGES = ["1h", "24h", "7d", "30d", "all"]
VALID_BUCKETS = ["hour", "day"]


def since_ts(range_: str) -> str | None:
    if range_ == "all":
        return None
    return (datetime.now(timezone.utc) - _RANGE_DELTAS.get(range_, timedelta(days=30))).strftime("%Y-%m-%dT%H:%M:%SZ")


def _percentile(values, p):
    if not values:
        return None
    s = sorted(values)
    k = max(0, min(len(s) - 1, int(round((p / 100) * (len(s) - 1)))))
    return s[k]


def _mask(v):
    if not v:
        return "пусто"
    return f"установлен (…{v[-4:]})"


async def get_summary(range_):
    since = since_ts(range_)
    base = await db.get_summary(since)
    lat = await db.get_latencies(since)
    total = base.get("total_requests") or 0
    errs = base.get("error_count") or 0
    base["error_rate"] = round(errs / total, 4) if total else 0.0
    base["p50_latency_ms"] = _percentile(lat, 50)
    base["p95_latency_ms"] = _percentile(lat, 95)
    base["total_cost"] = round(base.get("total_cost") or 0.0, 6)
    base["avg_latency_ms"] = round(base.get("avg_latency_ms") or 0.0, 1)
    return base


async def get_timeline(range_, bucket): return await db.get_timeline(since_ts(range_), bucket)


async def get_models(range_):
    rows = await db.get_models(since_ts(range_))
    for r in rows:
        reqs = r.get("requests") or 0
        r["error_rate"] = round((r.get("error_count") or 0) / reqs, 4) if reqs else 0.0
        r["total_cost"] = round(r.get("total_cost") or 0.0, 6)
        r["avg_latency_ms"] = round(r.get("avg_latency_ms") or 0.0, 1)
    return rows


async def get_emails(range_):
    rows = await db.get_emails(since_ts(range_))
    balances = await db.get_email_balances()
    for r in rows:
        r["total_cost"] = round(r.get("total_cost") or 0.0, 6)
        b = balances.get(r["email"]) or {}
        r["balance"] = b.get("balance")
        r["charged"] = b.get("charged")
        r["balance_ts"] = b.get("balance_ts")
    return rows


async def get_user_report(range_):
    """Сводка по пользователям за период — основа Excel-выгрузки / акта работ."""
    rows = await db.get_user_report(since_ts(range_))
    for r in rows:
        r["total_cost"] = round(r.get("total_cost") or 0.0, 6)
        r["charged"] = round(r.get("charged"), 6) if r.get("charged") is not None else None
    return rows


async def get_requests(limit, offset, status, model, email, range_):
    return await db.get_requests(limit, offset, status, model, email, since_ts(range_))


def build_report_xlsx(range_: str, email: str | None) -> bytes:
    """Excel-отчёт за период: сводка по пользователям (для акта выполненных работ),
    детализация запросов и лист с итогами. Синхронный — вызывается из to_thread."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    since = since_ts(range_)
    users = asyncio.run(get_user_report(range_))
    if email:
        users = [u for u in users if u["email"] == email]
    reqs = asyncio.run(get_requests(5000, 0, None, None, email, range_))["items"]
    balances = asyncio.run(db.get_email_balances())

    wb = Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDE6F0")

    # Лист 1: сводка по пользователям
    ws = wb.active
    ws.title = "Сводка по пользователям"
    headers = ["Email", "Запросов", "Успешно", "Ошибок", "Стоимость, $",
               "Токенов (вх/вых)", "Всего токенов", "Картинок", "Период активности",
               "Списано внешним сервисом, $", "Баланс, $", "Баланс на момент"]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for u in users:
        ws.append([u["email"], u["requests"], u["success_count"], u["error_count"],
                   u["total_cost"], f'{u["prompt_tokens"] or 0} / {u["completion_tokens"] or 0}',
                   u["total_tokens"], u["total_images"], f'{u["first_ts"]} — {u["last_ts"]}',
                   u.get("charged"), u.get("balance"), u.get("balance_ts")])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.column_dimensions["A"].width = 28
    for col, w in (("J", 16), ("L", 22), ("I", 42)):
        ws.column_dimensions[col].width = w
    r0 = ws.max_row + 2
    ws.cell(row=r0, column=1, value="ИТОГО стоимость за период, $").font = bold
    ws.cell(row=r0, column=5, value=round(sum(u["total_cost"] or 0 for u in users), 6)).font = bold
    ws.cell(row=r0 + 1, column=1, value="Период выгрузки").font = bold
    ws.cell(row=r0 + 1, column=5, value=f"{range_} (с {since or 'начала наблюдений'})")
    ws.cell(row=r0 + 2, column=1, value="Сформировано").font = bold
    ws.cell(row=r0 + 2, column=5,
            value=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"))

    # Лист 2: детализация запросов
    ws2 = wb.create_sheet("Запросы")
    h2 = ["Время (UTC)", "Email", "Task", "Модель (запрошена)", "Модель (факт)",
          "Ток. вх", "Ток. вых", "Всего ток.", "Кэш ток.", "Reasoning ток.",
          "Стоимость, $", "Задержка, мс", "Статус", "Ошибка"]
    ws2.append(h2)
    for c in ws2[1]:
        c.font = bold
        c.fill = head_fill
    for it in reqs:
        ws2.append([it["ts"], it["email"], it["task"], it["model_requested"],
                    it["model_actual"], it["prompt_tokens"], it["completion_tokens"],
                    it["total_tokens"], it["cached_tokens"], it["reasoning_tokens"],
                    it["cost"], it["latency_ms"], it["status"],
                    (it["error"] or "")[:300]])
    for i in range(1, len(h2) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 18
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["N"].width = 40

    # Лист 3: балансы пользователей (последние известные)
    ws3 = wb.create_sheet("Балансы")
    ws3.append(["Email", "Баланс, $", "Списано всего, $", "Баланс на момент"])
    for c in ws3[1]:
        c.font = bold
        c.fill = head_fill
    for mail, b in sorted(balances.items()):
        ws3.append([mail, b.get("balance"), b.get("charged"), b.get("balance_ts")])
    for i, w in enumerate((28, 14, 18, 24), start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def get_external_summary(range_):
    d = await db.get_external_summary(since_ts(range_))
    d["total_charged"] = round(d.get("total_charged") or 0.0, 6)
    reports = d.get("reports") or 0
    d["send_success_rate"] = round((d.get("sent_ok") or 0) / reports, 4) if reports else 0.0
    return d


async def get_external_reports(limit, offset, range_):
    return await db.get_external_reports(limit, offset, since_ts(range_))


# Кэш живой проверки внешнего сервиса: {ts, result}. TTL защищает от спама при частых опросах.
_probe_cache = {"ts": 0.0, "result": None}
_PROBE_TTL = 30.0


async def probe_external():
    """Живая проверка доступности внешнего сервиса (webhook/billing).

    GET к settings.webhook_url: любой HTTP-ответ — сервис доступен,
    ошибка соединения/таймаут — недоступен. Результат кэшируется на _PROBE_TTL сек,
    чтобы частый опрос дашборда не спамил внешний сервис.

    Адаптация (Агент Норм): биллинг-webhook не используется — при выключенном
    webhook и заданном TUNNEL_PUBLIC_HOST индикатор «Внешний» проверяет
    публичный доступ к сервису через туннель (GET https://<host>/health).
    Зелёный «Доступен» = сервис реально отвечает из интернета.
    """
    if settings.webhook_enabled:
        url = settings.webhook_url
    else:
        host = (settings.tunnel_public_host or "").strip()
        if not host:
            return {"enabled": False, "reachable": None, "status": None, "latency_ms": None,
                    "error": "webhook выключен и TUNNEL_PUBLIC_HOST не задан", "url": "",
                    "cached": False}
        url = f"https://{host}/health"

    now = time.monotonic()
    cached = _probe_cache["result"]
    if (cached is not None and (now - _probe_cache["ts"]) < _PROBE_TTL
            and cached.get("url") == url):
        out = dict(cached)
        out["cached"] = True
        return out

    t0 = time.perf_counter()
    try:
        async with httpx.AsyncClient(timeout=httpx.Timeout(6.0, connect=5.0)) as client:
            resp = await client.get(url)
        latency_ms = int((time.perf_counter() - t0) * 1000)
        result = {"enabled": True, "reachable": True, "status": resp.status_code,
                  "latency_ms": latency_ms, "error": None, "url": url}
    except Exception as e:
        latency_ms = int((time.perf_counter() - t0) * 1000)
        result = {"enabled": True, "reachable": False, "status": None, "latency_ms": latency_ms,
                  "error": f"{type(e).__name__}: {str(e)[:200]}", "url": url}

    _probe_cache["ts"] = now
    _probe_cache["result"] = result
    out = dict(result)
    out["cached"] = False
    return out


async def get_credits(live):
    data = {"snapshot": await db.get_latest_credit()}
    if live:
        data["live"] = await fetch_openrouter_credits()
        snap = data["live"]
        if snap is not None:
            await db.save_credits(snap.get("total_credits"), snap.get("total_usage"),
                                  snap.get("remaining"), json.dumps(snap.get("raw") or {}, ensure_ascii=False))
            data["snapshot"] = await db.get_latest_credit()
    return data


def get_settings():
    return {
        "server": {"app_title": settings.analytics_app_title, "openrouter_base_url": settings.openrouter_base_url,
                   "default_model": settings.default_model, "models": settings.models_list,
                   "api_key": _mask(settings.openrouter_api_key)},
        "analytics": {"track_usage": settings.analytics_track_usage,
                      "credits_poll_interval": settings.credits_poll_interval,
                      "db_path": settings.analytics_db_path, "require_auth": settings.analytics_require_auth},
        "webhook": {"enabled": settings.webhook_enabled, "url": settings.webhook_url,
                    "auth_header": settings.webhook_auth_header, "token": _mask(settings.webhook_token),
                    "user_id": _mask(settings.webhook_user_id) if settings.webhook_user_id else "не задан",
                    "timeout": settings.webhook_timeout},
        "tunnel": {"name": settings.tunnel_name, "public_host": settings.tunnel_public_host},
    }


def _auth_headers():
    return {"Authorization": f"Bearer {settings.openrouter_api_key}"}


async def fetch_openrouter_credits():
    if not settings.openrouter_api_key:
        return None
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(f"{settings.openrouter_base_url}/credits", headers=_auth_headers())
        if resp.status_code >= 400:
            return None
        d = resp.json().get("data") or {}
        total, usage = d.get("total_credits"), d.get("total_usage")
        remaining = (total - usage) if (total is not None and usage is not None) else None
        return {"total_credits": total, "total_usage": usage, "remaining": remaining, "raw": d}
    except Exception:
        log.exception("fetch credits failed"); return None


async def fetch_openrouter_key_info():
    if not settings.openrouter_api_key:
        return None
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(f"{settings.openrouter_base_url}/key", headers=_auth_headers())
        if resp.status_code >= 400:
            return None
        return resp.json().get("data") or {}
    except Exception:
        log.exception("fetch key failed"); return None


async def credits_poll_loop():
    interval = max(60, settings.credits_poll_interval)
    while True:
        snap = await fetch_openrouter_credits()
        if snap is not None:
            await db.save_credits(snap.get("total_credits"), snap.get("total_usage"),
                                  snap.get("remaining"), json.dumps(snap.get("raw") or {}, ensure_ascii=False))
        await asyncio.sleep(interval)
