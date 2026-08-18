"""FastAPI-сервер списания кредитов.

Принимает от сниппетов (в Open WebUI функциях) данные об использовании:
email, функция, модель, дата/время, токены — считает кредиты по курсу модели,
хранит в SQLite и пересылает на «другой сервер».

Запуск: uvicorn server:app --host 0.0.0.0 --port 4010
GUI:   http://localhost:4010/admin
"""
from __future__ import annotations

import hmac
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import (
    FastAPI,
    HTTPException,
    Query,
    Request,
)
from fastapi.responses import HTMLResponse, Response
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field

import converter
import db
import forwarder
from analytics import (
    router as analytics_router,
    init_analytics,
    install_log_buffer,
    tracker,
)


# --------------------------------------------------------------------------- #
# .env (без python-dotenv) — повторяет паттерн doc_server/server.py
# --------------------------------------------------------------------------- #
def _load_env(path: str | Path) -> None:
    p = Path(path)
    if not p.is_file():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        key = key.strip()
        val = val.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = val


_load_env(Path(__file__).resolve().parent / ".env")

# --------------------------------------------------------------------------- #
# Конфигурация
# --------------------------------------------------------------------------- #
DB_PATH = os.environ.get("DB_PATH", "credits.db")
BASE_URL = os.environ.get("BASE_URL", "http://localhost:4010").rstrip("/")

# Bearer приёма данных (/api/usage). Пустой = auth выключен (только dev).
CREDITS_API_KEY = os.environ.get("CREDITS_API_KEY", "")
# Bearer админки (/admin и /api/rates|stats|records|settings|retry-failed).
ADMIN_KEY = os.environ.get("ADMIN_KEY", "")


# Логгер для дашборда аналитики (страница «Логи» ловит logging, не stdout).
_srv_log = logging.getLogger("server")
_srv_log.setLevel(logging.INFO)


def _log(msg: str) -> None:
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%SZ")
    print(f"[server] {ts} {msg}", file=sys.stdout, flush=True)
    _srv_log.info(msg)


def _check_bearer(request: Request, expected: str, *, what: str) -> None:
    """Constant-time проверка Bearer. Пустой expected = auth выключен."""
    if not expected:
        return  # auth выключен — локальная разработка
    auth = request.headers.get("Authorization", "")
    token = auth[len("Bearer "):] if auth.startswith("Bearer ") else ""
    if not hmac.compare_digest(token, expected):
        raise HTTPException(status_code=401, detail=f"Invalid or missing {what} key")


# --------------------------------------------------------------------------- #
# Pydantic-модели
# --------------------------------------------------------------------------- #
class UsageIn(BaseModel):
    email: str = Field(..., description="Email пользователя Open WebUI")
    function: str = Field(..., description="Название функции (агента)")
    model: str = Field("", description="Имя модели (для выбора курса). Может быть пустым.")
    timestamp: str = Field(
        "",
        description="Дата/время операции (ISO). Если пусто — берётся now(UTC).",
    )
    tokens: int = Field(..., ge=0, description="Всего токенов (>= 0)")
    prompt_tokens: Optional[int] = Field(
        default=None, ge=0,
        description="Breakdown: токены input. Если задано вместе с "
        "completion_tokens и в курсе модели есть input/output — применяется "
        "раздельная тарификация.",
    )
    completion_tokens: Optional[int] = Field(
        default=None, ge=0, description="Breakdown: токены output (см. prompt_tokens)."
    )
    cost_usd: Optional[float] = Field(
        default=None, ge=0,
        description="Стоимость в USD от провайдера (напр. usage.cost у OpenRouter). "
        "Если нет — сервер посчитает по тарифам $/млн из курса модели.",
    )
    user_id: str = Field(
        "",
        description="Внутренний ID пользователя Open WebUI (openRouterWebUiUserId). "
        "Пробрасывается в пересылку на devbim.com.",
    )
    request_date: Optional[str] = Field(
        default=None,
        description="ISO-время старта запроса к модели. Пробрасывается в пересылку "
        "(requestDate контракта devbim.com).",
    )
    response_date: Optional[str] = Field(
        default=None,
        description="ISO-время получения ответа модели. Пробрасывается в пересылку "
        "(responseDate контракта devbim.com).",
    )
    is_success: bool = Field(
        True,
        description="Флаг успешного вызова модели (isSuccess контракта devbim.com). "
        "False — модель вернула ошибку/недоступна.",
    )
    error_message: Optional[str] = Field(
        default=None,
        description="Текст ошибки модели (errorMessage контракта devbim.com). "
        "Заполняется только при is_success=False.",
    )


class UsageOut(BaseModel):
    ok: bool = True
    credits: float
    rate: float
    matched_model: str
    rate_input: Optional[float] = None
    rate_output: Optional[float] = None
    pricing: str = "base"  # "base" или "split"
    cost_usd: Optional[float] = None
    cost_source: Optional[str] = None  # "provider" | "computed" | None


class RateIn(BaseModel):
    model_name: str = Field(..., description="Имя модели или __default__")
    credits_per_1k_tokens: float = Field(..., ge=0, description="Единый курс за 1000 токенов")
    credits_per_1k_input: Optional[float] = Field(
        default=None, ge=0,
        description="Раздельный курс за 1000 input-токенов. NULL/пусто = выключено.",
    )
    credits_per_1k_output: Optional[float] = Field(
        default=None, ge=0,
        description="Раздельный курс за 1000 output-токенов. NULL/пусто = выключено.",
    )
    cost_per_1m_input_usd: Optional[float] = Field(
        default=None, ge=0,
        description="Цена за 1М input-токенов в $. Нужна для расчёта стоимости "
        "провайдеров без поля cost (z.ai/GLM). NULL/пусто = не считать $.",
    )
    cost_per_1m_output_usd: Optional[float] = Field(
        default=None, ge=0,
        description="Цена за 1М output-токенов в $ (см. cost_per_1m_input_usd).",
    )


class SettingsIn(BaseModel):
    forward_url: str = ""
    forward_api_key: str = ""


def _report_latency_ms(req: UsageIn) -> Optional[int]:
    """Latency отчёта из request_date/response_date (если пайп передал оба)."""
    if not (req.request_date and req.response_date):
        return None
    try:
        t0 = datetime.fromisoformat(req.request_date.replace("Z", "+00:00"))
        t1 = datetime.fromisoformat(req.response_date.replace("Z", "+00:00"))
    except ValueError:
        return None
    return max(0, int((t1 - t0).total_seconds() * 1000))


# --------------------------------------------------------------------------- #
# App
# --------------------------------------------------------------------------- #
app = FastAPI(title="Сервер списания кредитов", version="1.0.0")
templates = Jinja2Templates(directory=str(Path(__file__).resolve().parent / "templates"))

# Дашборд аналитики (/dashboard + /stats/*) — самодостаточный пакет analytics/.
app.include_router(analytics_router)


@app.on_event("startup")
async def _on_startup() -> None:
    install_log_buffer()  # захват логов для страницы «Логи» дашборда
    await init_analytics()  # БД аналитики + опрос баланса (если задан ключ)
    # .env разворачивает стартовую конфигурацию пересылки. Если в GUI уже задано
    # непустое значение — оно сохранится (init_db перетирает только пустые поля).
    env_overrides = {
        "forward_url": os.environ.get("FORWARD_URL", "").strip(),
        "forward_api_key": os.environ.get("FORWARD_API_KEY", "").strip(),
    }
    await db.init_db(DB_PATH, env_overrides=env_overrides)
    _log(f"init OK db={DB_PATH!r} base={BASE_URL!r}")


# --------------------------------------------------------------------------- #
# Корень → редирект на админку (чтобы http://localhost:4010/ открывал GUI)
# --------------------------------------------------------------------------- #
from fastapi.responses import RedirectResponse


@app.get("/", include_in_schema=False)
def root():
    return RedirectResponse(url="/admin", status_code=302)


# --------------------------------------------------------------------------- #
# Health
# --------------------------------------------------------------------------- #
@app.get("/health")
def health():
    return {"status": "ok"}


# --------------------------------------------------------------------------- #
# Приём данных от сниппета
# --------------------------------------------------------------------------- #
@app.post("/api/usage", response_model=UsageOut)
async def post_usage(req: UsageIn, request: Request):
    _check_bearer(request, CREDITS_API_KEY, what="CREDITS_API_KEY")

    op_ts = req.timestamp.strip() or datetime.now(timezone.utc).isoformat(timespec="seconds")
    model = (req.model or "").strip()

    credits, rate, matched, rate_in, rate_out = await converter.compute_credits(
        DB_PATH,
        model,
        req.tokens,
        prompt_tokens=req.prompt_tokens,
        completion_tokens=req.completion_tokens,
    )
    pricing = "split" if (rate_in is not None and rate_out is not None) else "base"

    # Стоимость в долларах: от провайдера (если передал) или расчёт по тарифам $/млн.
    cost_usd, cost_source = await converter.compute_cost_usd(
        DB_PATH,
        model,
        req.prompt_tokens,
        req.completion_tokens,
        provider_cost=req.cost_usd,
    )

    record_id = await db.insert_usage(
        DB_PATH,
        email=req.email.strip(),
        function=req.function.strip(),
        model=model or db.DEFAULT_RATE_KEY,
        op_timestamp=op_ts,
        tokens=req.tokens,
        rate_applied=rate,
        credits=credits,
        prompt_tokens=req.prompt_tokens,
        completion_tokens=req.completion_tokens,
        rate_input=rate_in,
        rate_output=rate_out,
        cost_usd=cost_usd,
        cost_source=cost_source,
        user_id=(req.user_id or "").strip() or None,
        request_date=(req.request_date or "").strip() or None,
        response_date=(req.response_date or "").strip() or None,
        is_success=req.is_success,
        error_message=(req.error_message or "").strip() or None,
    )
    # Создаём pending-запись и планируем пересылку (fire-and-forget).
    await db.create_forward_entry(DB_PATH, record_id)
    forwarder.schedule_forward(DB_PATH)

    # Аналитика дашборда (/dashboard): пишем копию отчёта в analytics.db.
    # Сами LLM не вызываем — точкой трекинга служит приём отчёта от пайпа.
    # Webhook аналитики выключен: пересылку на devbim.com делает forwarder выше
    # (иначе было бы двойное списание).
    await tracker.track(
        model_requested=model or db.DEFAULT_RATE_KEY,
        model_actual=model or db.DEFAULT_RATE_KEY,
        usage={
            "prompt_tokens": req.prompt_tokens,
            "completion_tokens": req.completion_tokens,
            "total_tokens": req.tokens,
            "cost": req.cost_usd,
        },
        email=req.email.strip(),
        user_id=(req.user_id or "").strip() or None,
        latency_ms=_report_latency_ms(req),
        status="success" if req.is_success else "error",
        error=(req.error_message or "").strip() or None,
        task=req.function.strip(),  # имя функции (агента) Open WebUI
        request_date=req.request_date,
        response_date=req.response_date,
    )

    _log(
        f"USAGE id={record_id} email={req.email!r} func={req.function!r} "
        f"model={model!r} tokens={req.tokens} "
        f"(in={req.prompt_tokens}/out={req.completion_tokens}) "
        f"pricing={pricing} credits={credits} rate={rate}"
        + (f" in={rate_in}/out={rate_out}" if pricing == "split" else "")
        + (f" cost=${cost_usd}({cost_source})" if cost_usd is not None else "")
    )
    return UsageOut(
        ok=True, credits=credits, rate=rate, matched_model=matched,
        rate_input=rate_in, rate_output=rate_out, pricing=pricing,
        cost_usd=cost_usd, cost_source=cost_source,
    )


# --------------------------------------------------------------------------- #
# GUI: страница
# --------------------------------------------------------------------------- #
@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    # Саму страницу отдаём без проверки ключа — ключ проверяется в JS на каждый
    # запрос к /api/*. Это упрощает первоначальное открытие и хранение ключа в
    # localStorage, как условлено в плане.
    return templates.TemplateResponse(
        "admin.html",
        {"request": request, "base_url": BASE_URL, "has_admin_key": bool(ADMIN_KEY)},
    )


# --------------------------------------------------------------------------- #
# GUI: страница отчётов
# --------------------------------------------------------------------------- #
@app.get("/report", response_class=HTMLResponse)
async def report_page(request: Request):
    # Как и /admin: страница без ключа, ключ (ADMIN_KEY) проверяется в JS на /api/*.
    return templates.TemplateResponse(
        "report.html",
        {"request": request, "base_url": BASE_URL, "has_admin_key": bool(ADMIN_KEY)},
    )


# --------------------------------------------------------------------------- #
# GUI API: курсы
# --------------------------------------------------------------------------- #
@app.get("/api/rates")
async def api_list_rates(request: Request):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    return await db.list_rates(DB_PATH)


@app.post("/api/rates")
async def api_upsert_rate(rate: RateIn, request: Request):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    name = rate.model_name.strip()
    if not name:
        raise HTTPException(status_code=422, detail="model_name required")
    await db.upsert_rate(
        DB_PATH,
        name,
        rate.credits_per_1k_tokens,
        credits_per_1k_input=rate.credits_per_1k_input,
        credits_per_1k_output=rate.credits_per_1k_output,
        cost_per_1m_input_usd=rate.cost_per_1m_input_usd,
        cost_per_1m_output_usd=rate.cost_per_1m_output_usd,
    )
    return {"ok": True}


@app.delete("/api/rates/{model_name}")
async def api_delete_rate(model_name: str, request: Request):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    try:
        await db.delete_rate(DB_PATH, model_name)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"ok": True}


# --------------------------------------------------------------------------- #
# GUI API: записи и статистика
# --------------------------------------------------------------------------- #
@app.get("/api/records")
async def api_list_records(
    request: Request,
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    email: Optional[str] = Query(None),
    function: Optional[str] = Query(None),
):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    rows, total = await db.list_records(
        DB_PATH, limit=limit, offset=offset, email=email, function=function
    )
    return {"rows": rows, "total": total, "limit": limit, "offset": offset}


@app.get("/api/stats")
async def api_stats(
    request: Request,
    group_by: str = Query("email", pattern="^(email|function|model)$"),
):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    return {"group_by": group_by, "rows": await db.get_stats(DB_PATH, group_by=group_by)}


@app.get("/api/summary")
async def api_summary(request: Request):
    """Сводные итоги для шапки админки: всего $, кредитов, вызовов."""
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    return await db.get_summary(DB_PATH)


# --------------------------------------------------------------------------- #
# GUI API: очередь пересылки
# --------------------------------------------------------------------------- #
@app.get("/api/forward")
async def api_list_forward(
    request: Request,
    status: Optional[str] = Query(None, pattern="^(pending|ok|failed)$"),
    limit: int = Query(50, ge=1, le=500),
):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    return {"rows": await db.list_forward(DB_PATH, status=status, limit=limit)}


@app.post("/api/retry-failed")
async def api_retry_failed(request: Request):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    n = await forwarder.retry_failed(DB_PATH)
    return {"ok": True, "retried": n}


# --------------------------------------------------------------------------- #
# GUI API: реестр балансов
# --------------------------------------------------------------------------- #
@app.get("/api/balances")
async def api_list_balances(request: Request):
    """Актуальные балансы пользователей из ответов внешнего сервера."""
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    return {"rows": await db.list_balances(DB_PATH)}


# --------------------------------------------------------------------------- #
# GUI API: настройки
# --------------------------------------------------------------------------- #
@app.get("/api/settings")
async def api_get_settings(request: Request):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    s = await db.get_all_settings(DB_PATH)
    # forward_api_key не возвращаем целиком из соображений безопасности.
    return {
        "forward_url": s.get("forward_url", ""),
        "has_forward_api_key": bool(s.get("forward_api_key", "")),
    }


@app.post("/api/settings")
async def api_set_settings(settings: SettingsIn, request: Request):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    await db.set_setting(DB_PATH, "forward_url", settings.forward_url.strip())
    # Пустой ключ в запросе = не менять существующий.
    if settings.forward_api_key.strip():
        await db.set_setting(DB_PATH, "forward_api_key", settings.forward_api_key.strip())
    return {"ok": True}


# --------------------------------------------------------------------------- #
# GUI API: отчёты (по пользователям / за период, с экспортом)
# --------------------------------------------------------------------------- #
def _report_filters(
    email: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
    group_by: str,
) -> dict:
    return {
        "email": (email or "").strip() or None,
        "date_from": date_from,
        "date_to": date_to,
        "group_by": group_by,
    }


@app.get("/api/report/emails")
async def api_report_emails(request: Request):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    return {"emails": await db.list_report_emails(DB_PATH)}


@app.get("/api/report")
async def api_report(
    request: Request,
    email: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    group_by: str = Query("month", pattern="^(month|day|email)$"),
):
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    params = _report_filters(email, date_from, date_to, group_by)
    rows = await db.get_report(DB_PATH, **params)
    totals = {
        "requests": sum(r["requests"] for r in rows),
        "ok_requests": sum(r["ok_requests"] for r in rows),
        "tokens": sum(r["tokens"] or 0 for r in rows),
        "prompt_tokens": sum(r["prompt_tokens"] or 0 for r in rows),
        "completion_tokens": sum(r["completion_tokens"] or 0 for r in rows),
        "credits": round(sum(r["credits"] or 0 for r in rows), 2),
        "cost_usd": round(sum(r["cost_usd"] or 0 for r in rows), 6),
    }
    return {"params": params, "rows": rows, "totals": totals}


@app.get("/api/report/export")
async def api_report_export(
    request: Request,
    email: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    group_by: str = Query("month", pattern="^(month|day|email)$"),
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
):
    """Отчёт файлом: format=xlsx (Excel, openpyxl) или csv. Фильтры те же."""
    _check_bearer(request, ADMIN_KEY, what="ADMIN")
    params = _report_filters(email, date_from, date_to, group_by)
    rows = await db.get_report(DB_PATH, **params)

    headers = ["Период" if params["group_by"] != "email" else "Email",
               "Запросов", "Успешных", "Токены", "Токены input", "Токены output",
               "Кредиты", "Стоимость $"]
    table = [[r["period"], r["requests"], r["ok_requests"], r["tokens"],
              r["prompt_tokens"], r["completion_tokens"], r["credits"],
              r["cost_usd"]] for r in rows]

    period_lbl = "по всем пользователям" if not params["email"] else params["email"]
    range_lbl = (f"{params['date_from'] or '…'} — {params['date_to'] or '…'}")
    fname_base = f"report_{params['email'] or 'all'}_{params['group_by']}"

    if format == "csv":
        import csv
        import io
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        w.writerows(table)
        return Response(
            content="\ufeff" + buf.getvalue(),  # BOM: Excel открывает UTF-8 корректно
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{fname_base}.csv"'},
        )

    # xlsx по умолчанию
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    ws.append(["Отчёт использования", period_lbl, range_lbl])
    ws["A1"].font = Font(bold=True)
    ws.append(headers)
    for c in ws[2]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
    for row in table:
        ws.append(row)
    widths = [22, 10, 10, 12, 13, 14, 10, 12]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = wd
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname_base}.xlsx"'},
    )


# --------------------------------------------------------------------------- #
# Запуск как скрипт
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    import uvicorn

    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "4010"))
    uvicorn.run("server:app", host=host, port=port, reload=False)
